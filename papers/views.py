import random
import openpyxl
from django.http import HttpResponse
from django.template.loader import render_to_string
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from .models import Subject, Question, GeneratedPaper
from .serializers import (SubjectSerializer, QuestionSerializer,
                           GeneratedPaperSerializer, PaperListSerializer)

# CO selection rule: 2 from CO1, 2 from CO2, 1 from CO3
CO_SELECTION = {'CO1': 2, 'CO2': 2, 'CO3': 1}

def select_questions_for_part(subject, part):
    selected = []
    for co, count in CO_SELECTION.items():
        pool = list(Question.objects.filter(subject=subject, part=part, co=co))
        if len(pool) < count:
            raise ValueError(
                f"Part-{part} / {co}: only {len(pool)} question(s) available, need {count}."
            )
        selected.extend(random.sample(pool, count))
    return selected


# ── Subjects ──────────────────────────────────────────────────────────────────

class SubjectListCreateView(APIView):
    def get(self, request):
        subjects = Subject.objects.all().order_by('name')
        return Response(SubjectSerializer(subjects, many=True).data)

    def post(self, request):
        s = SubjectSerializer(data=request.data)
        if s.is_valid():
            s.save()
            return Response(s.data, status=status.HTTP_201_CREATED)
        return Response(s.errors, status=status.HTTP_400_BAD_REQUEST)

class SubjectDeleteView(APIView):
    def delete(self, request, pk):
        try:
            Subject.objects.get(pk=pk).delete()
            return Response({'message': 'Deleted'}, status=status.HTTP_204_NO_CONTENT)
        except Subject.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)


# ── Questions ─────────────────────────────────────────────────────────────────

class QuestionListView(APIView):
    def get(self, request):
        subject_id = request.query_params.get('subject')
        part       = request.query_params.get('part')
        co         = request.query_params.get('co')
        qs = Question.objects.select_related('subject').all()
        if subject_id: qs = qs.filter(subject_id=subject_id)
        if part:       qs = qs.filter(part=part.upper())
        if co:         qs = qs.filter(co=co.upper())
        return Response(QuestionSerializer(qs, many=True).data)

class QuestionDeleteView(APIView):
    def delete(self, request, pk):
        try:
            Question.objects.get(pk=pk).delete()
            return Response({'message': 'Deleted'}, status=status.HTTP_204_NO_CONTENT)
        except Question.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)


# ── Excel Upload ──────────────────────────────────────────────────────────────

class ExcelUploadView(APIView):
    def post(self, request):
        file       = request.FILES.get('file')
        subject_id = request.data.get('subject_id')

        if not file or not subject_id:
            return Response({'error': 'file and subject_id are required'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            subject = Subject.objects.get(pk=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'},
                            status=status.HTTP_404_NOT_FOUND)
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception:
            return Response({'error': 'Invalid Excel file'},
                            status=status.HTTP_400_BAD_REQUEST)

        questions, errors = [], []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # cols: 0=Q.No | 1=Part | 2=CO | 3=BTL | 4=Question Text
            if not row[4]:
                continue
            try:
                part = str(row[1]).strip().upper()
                co   = str(row[2]).strip().upper()
                btl  = int(row[3])
                text = str(row[4]).strip()

                if part not in ['A', 'B', 'C']:
                    raise ValueError(f"Invalid Part '{part}'")
                if co not in ['CO1', 'CO2', 'CO3']:
                    raise ValueError(f"Invalid CO '{co}'")
                if btl not in range(1, 7):
                    raise ValueError(f"Invalid BTL '{btl}'")

                questions.append(Question(
                    subject=subject, part=part, co=co,
                    btl=btl, text=text,
                    uploaded_by=request.user,
                ))
            except Exception as e:
                errors.append(f"Row {i}: {str(e)}")

        Question.objects.bulk_create(questions)
        return Response({
            'created': len(questions),
            'errors':  errors,
            'message': f'{len(questions)} questions uploaded successfully.'
        }, status=status.HTTP_201_CREATED)


# ── Generate Paper ────────────────────────────────────────────────────────────

class GeneratePaperView(APIView):
    def post(self, request):
        data       = request.data
        subject_id = data.get('subject_id')

        try:
            subject = Subject.objects.get(pk=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'},
                            status=status.HTTP_404_NOT_FOUND)

        all_selected = []
        for part in ['A', 'B', 'C']:
            try:
                all_selected.extend(select_questions_for_part(subject, part))
            except ValueError as e:
                return Response({'error': str(e)},
                                status=status.HTTP_400_BAD_REQUEST)

        paper = GeneratedPaper.objects.create(
            subject      = subject,
            title        = data.get('title', 'Mid Term Examination'),
            program      = data.get('course', ''),
            semester     = data.get('semester', ''),
            course_name  = data.get('course_name', subject.name),
            course_code  = data.get('course_code', subject.code),
            exam_type    = data.get('exam_type', 'MID TERM EXAMINATION'),
            session      = data.get('session', '2025-26'),
            duration     = data.get('duration', '3 Hours'),
            max_marks    = int(data.get('max_marks', 100)),
            instructions = data.get('instructions',
                'Attempt all parts given in the question paper. '
                'No student is allowed to leave the examination hall '
                'before the completion of the time.'),
            created_by   = request.user,
        )
        paper.questions.set(all_selected)
        return Response(GeneratedPaperSerializer(paper).data,
                        status=status.HTTP_201_CREATED)


# ── Paper List & Detail ───────────────────────────────────────────────────────

class PaperListView(APIView):
    def get(self, request):
        papers = GeneratedPaper.objects.select_related('subject', 'created_by')\
                                       .order_by('-created_at')
        return Response(PaperListSerializer(papers, many=True).data)

class PaperDetailView(APIView):
    def get(self, request, pk):
        try:
            paper = GeneratedPaper.objects.prefetch_related('questions').get(pk=pk)
            return Response(GeneratedPaperSerializer(paper).data)
        except GeneratedPaper.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

class PaperDeleteView(APIView):
    def delete(self, request, pk):
        try:
            GeneratedPaper.objects.get(pk=pk).delete()
            return Response({'message': 'Deleted'}, status=status.HTTP_204_NO_CONTENT)
        except GeneratedPaper.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)


# ── Download PDF (HTML for now) ───────────────────────────────────────────────

class DownloadPDFView(APIView):
    def get(self, request, pk):
        try:
            paper = GeneratedPaper.objects.prefetch_related('questions')\
                                          .select_related('subject').get(pk=pk)
        except GeneratedPaper.DoesNotExist:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)

        import re
        def format_question_text(q):
            text = q.text.replace('\r', '')
            # Remove "Answer: ..." from the bottom
            text = re.sub(r'(?i)\n*ans(?:wer)?\s*[:\-].*$', '', text)
            # Add newline before options (a), (b), (c), (d) if not already there
            text = re.sub(r'(?<!\n)\s*(\([a-d]\)|[a-d]\)) ', r'\n\1 ', text)
            q.formatted_text = text.strip()
            return q

        part_a = [format_question_text(q) for q in paper.questions.filter(part='A').order_by('co')]
        part_b = [format_question_text(q) for q in paper.questions.filter(part='B').order_by('co')]
        part_c = [format_question_text(q) for q in paper.questions.filter(part='C').order_by('co')]

        import os
        from django.conf import settings
        logo_path = os.path.join(settings.BASE_DIR, 'frontend', 'public', 'coer-logo.png')

        html = render_to_string('papers/paper_template.html', {
            'paper':  paper,
            'part_a': part_a,
            'part_b': part_b,
            'part_c': part_c,
            'logo_path': logo_path,
        })

        # Convert HTML → PDF using xhtml2pdf
        from io import BytesIO
        from xhtml2pdf import pisa

        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)

        if pisa_status.err:
            return Response({'error': 'PDF generation failed'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        pdf_buffer.seek(0)
        filename = f"QP_{paper.subject.name.replace(' ', '_')}.pdf"
        response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ── Excel Template Download ───────────────────────────────────────────────────

class ExcelTemplateView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        wb = openpyxl.Workbook()
        for part, label in [('A', 'Part-A'), ('B', 'Part-B'), ('C', 'Part-C')]:
            ws = wb.create_sheet(title=label)
            ws.append(['Q.No', 'Part', 'CO', 'BTL', 'Question Text'])
            co_cycle = ['CO1'] * 5 + ['CO2'] * 5 + ['CO3'] * 5
            for idx in range(15):
                ws.append([idx + 1, part, co_cycle[idx], 2,
                            f'Sample question {idx+1} for {label} ({co_cycle[idx]})'])
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = \
            'attachment; filename="coer_question_bank_template.xlsx"'
        wb.save(response)
        return response


# ── Dashboard Stats ───────────────────────────────────────────────────────────

class DashboardStatsView(APIView):
    def get(self, request):
        import datetime
        from django.db.models import Count, Q
        
        q_stats = Question.objects.aggregate(
            total=Count('id'),
            part_a=Count('id', filter=Q(part='A')),
            part_b=Count('id', filter=Q(part='B')),
            part_c=Count('id', filter=Q(part='C')),
            co1=Count('id', filter=Q(co='CO1')),
            co2=Count('id', filter=Q(co='CO2')),
            co3=Count('id', filter=Q(co='CO3')),
        )
        
        p_stats = GeneratedPaper.objects.aggregate(
            total=Count('id'),
            this_month=Count('id', filter=Q(created_at__month=datetime.date.today().month))
        )

        return Response({
            'total_subjects':    Subject.objects.count(),
            'total_questions':   q_stats['total'],
            'part_a_count':      q_stats['part_a'],
            'part_b_count':      q_stats['part_b'],
            'part_c_count':      q_stats['part_c'],
            'co1_count':         q_stats['co1'],
            'co2_count':         q_stats['co2'],
            'co3_count':         q_stats['co3'],
            'total_papers':      p_stats['total'],
            'papers_this_month': p_stats['this_month'],
        })